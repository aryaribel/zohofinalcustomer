#Zoho Final
from django.shortcuts import render,redirect
from Register_Login.models import *
from Register_Login.views import logout
from django.contrib import messages
from django.conf import settings
from datetime import date
from datetime import datetime, timedelta
from Company_Staff.models import *
from django.db import models
from django.shortcuts import get_object_or_404
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from xhtml2pdf import pisa
from django.template.loader import get_template
from bs4 import BeautifulSoup
import io,os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from django.http import HttpResponse,HttpResponseRedirect
from io import BytesIO
from django.db.models import Max
from django.db.models import Q
from django.http import JsonResponse,HttpResponse,HttpResponseRedirect

# Create your views here.
from decimal import Decimal
from django.http import HttpResponseNotFound, JsonResponse

from django.core.mail import EmailMultiAlternatives


# -------------------------------Company section--------------------------------
# company dashboard
def company_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        # Calculate the date 20 days before the end date for payment term renew and 10 days before for trial period renew
        if dash_details.payment_term:
            reminder_date = dash_details.End_date - timedelta(days=20)
        else:
            reminder_date = dash_details.End_date - timedelta(days=10)
        current_date = date.today()
        alert_message = current_date >= reminder_date
        
        payment_request = True if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists() else False

        # Calculate the number of days between the reminder date and end date
        days_left = (dash_details.End_date - current_date).days
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'alert_message':alert_message,
            'days_left':days_left,
            'payment_request':payment_request,
        }
        return render(request, 'company/company_dash.html', context)
    else:
        return redirect('/')
    
    
# def company_dashboard(request):
#     if 'login_id' in request.session:
#         log_id = request.session['login_id']
#         if 'login_id' not in request.session:
#             return redirect('/')
#         log_details= LoginDetails.objects.get(id=log_id)
#         dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
#         allmodules= ZohoModules.objects.get(company=dash_details,status='New')

#         # Calculate the date 20 days before the end date for payment term renew
#         reminder_date = dash_details.End_date - timedelta(days=20)
#         current_date = date.today()
#         alert_message = current_date >= reminder_date
        
#         payment_request = True if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists() else False

#         # Calculate the number of days between the reminder date and end date
#         days_left = (dash_details.End_date - current_date).days
#         context = {
#             'details': dash_details,
#             'allmodules': allmodules,
#             'alert_message':alert_message,
#             'days_left':days_left,
#             'payment_request':payment_request,
#         }
#         return render(request, 'company/company_dash.html', context)
#     else:
#         return redirect('/')


# company staff request for login approval
def company_staff_request(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        staff_request=StaffDetails.objects.filter(company=dash_details.id, company_approval=0).order_by('-id')
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'requests':staff_request,
        }
        return render(request, 'company/staff_request.html', context)
    else:
        return redirect('/')

# company staff accept or reject
def staff_request_accept(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    staff.company_approval=1
    staff.save()
    return redirect('company_staff_request')

def staff_request_reject(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    login_details=LoginDetails.objects.get(id=staff.company.id)
    login_details.delete()
    staff.delete()
    return redirect('company_staff_request')


# All company staff view, cancel staff approval
def company_all_staff(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        all_staffs=StaffDetails.objects.filter(company=dash_details.id, company_approval=1).order_by('-id')
       
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'staffs':all_staffs,
        }
        return render(request, 'company/all_staff_view.html', context)
    else:
        return redirect('/')

def staff_approval_cancel(request, pk):
    """
    Sets the company approval status to 2 for the specified staff member, effectively canceling staff approval.

    This function is designed to be used for canceling staff approval, and the company approval value is set to 2.
    This can be useful for identifying resigned staff under the company in the future.

    """
    staff = StaffDetails.objects.get(id=pk)
    staff.company_approval = 2
    staff.save()
    return redirect('company_all_staff')


# company profile, profile edit
def company_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        terms=PaymentTerms.objects.all()
        payment_history=dash_details.previous_plans.all()

        # Calculate the date 20 days before the end date
        reminder_date = dash_details.End_date - timedelta(days=20)
        current_date = date.today()
        renew_button = current_date >= reminder_date

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'renew_button': renew_button,
            'terms':terms,
            'payment_history':payment_history,
        }
        return render(request, 'company/company_profile.html', context)
    else:
        return redirect('/')

def company_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_profile_editpage.html', context)
    else:
        return redirect('/')

def company_profile_basicdetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            messages.success(request,'Updated')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
    
def company_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                if LoginDetails.objects.filter(password=password).exists():
                    messages.error(request,'Use another password')
                    return redirect('company_profile_editpage')
                else:
                    log_details.password=password
                    log_details.save()

            messages.success(request,'Password Changed')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
       
def company_profile_companydetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            gstno = request.POST.get('gstno')
            profile_pic = request.FILES.get('image')

            # Update the CompanyDetails object with form data
            dash_details.company_name = request.POST.get('cname')
            dash_details.contact = request.POST.get('phone')
            dash_details.address = request.POST.get('address')
            dash_details.city = request.POST.get('city')
            dash_details.state = request.POST.get('state')
            dash_details.country = request.POST.get('country')
            dash_details.pincode = request.POST.get('pincode')
            dash_details.pan_number = request.POST.get('pannumber')

            if gstno:
                dash_details.gst_no = gstno

            if profile_pic:
                dash_details.profile_pic = profile_pic

            dash_details.save()

            messages.success(request, 'Updated')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/')    

# company modules editpage
def company_module_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_module_editpage.html', context)
    else:
        return redirect('/')

def company_module_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        # Check for any previous module update request
        if ZohoModules.objects.filter(company=dash_details,status='Pending').exists():
            messages.warning(request,'You have a pending update request, wait for approval or contact our support team for any help..?')
            return redirect('company_profile')
        if request.method == 'POST':
            # Retrieve values
            items = request.POST.get('items', 0)
            price_list = request.POST.get('price_list', 0)
            stock_adjustment = request.POST.get('stock_adjustment', 0)
            godown = request.POST.get('godown', 0)

            cash_in_hand = request.POST.get('cash_in_hand', 0)
            offline_banking = request.POST.get('offline_banking', 0)
            upi = request.POST.get('upi', 0)
            bank_holders = request.POST.get('bank_holders', 0)
            cheque = request.POST.get('cheque', 0)
            loan_account = request.POST.get('loan_account', 0)

            customers = request.POST.get('customers', 0)
            invoice = request.POST.get('invoice', 0)
            estimate = request.POST.get('estimate', 0)
            sales_order = request.POST.get('sales_order', 0)
            recurring_invoice = request.POST.get('recurring_invoice', 0)
            retainer_invoice = request.POST.get('retainer_invoice', 0)
            credit_note = request.POST.get('credit_note', 0)
            payment_received = request.POST.get('payment_received', 0)
            delivery_challan = request.POST.get('delivery_challan', 0)

            vendors = request.POST.get('vendors', 0)
            bills = request.POST.get('bills', 0)
            recurring_bills = request.POST.get('recurring_bills', 0)
            vendor_credit = request.POST.get('vendor_credit', 0)
            purchase_order = request.POST.get('purchase_order', 0)
            expenses = request.POST.get('expenses', 0)
            recurring_expenses = request.POST.get('recurring_expenses', 0)
            payment_made = request.POST.get('payment_made', 0)

            projects = request.POST.get('projects', 0)

            chart_of_accounts = request.POST.get('chart_of_accounts', 0)
            manual_journal = request.POST.get('manual_journal', 0)

            eway_bill = request.POST.get('ewaybill', 0)

            employees = request.POST.get('employees', 0)
            employees_loan = request.POST.get('employees_loan', 0)
            holiday = request.POST.get('holiday', 0)
            attendance = request.POST.get('attendance', 0)
            salary_details = request.POST.get('salary_details', 0)

            reports = request.POST.get('reports', 0)

            update_action=1
            status='Pending'

            # Create a new ZohoModules instance and save it to the database
            data = ZohoModules(
                company=dash_details,
                items=items, price_list=price_list, stock_adjustment=stock_adjustment, godown=godown,
                cash_in_hand=cash_in_hand, offline_banking=offline_banking, upi=upi, bank_holders=bank_holders,
                cheque=cheque, loan_account=loan_account,
                customers=customers, invoice=invoice, estimate=estimate, sales_order=sales_order,
                recurring_invoice=recurring_invoice, retainer_invoice=retainer_invoice, credit_note=credit_note,
                payment_received=payment_received, delivery_challan=delivery_challan,
                vendors=vendors, bills=bills, recurring_bills=recurring_bills, vendor_credit=vendor_credit,
                purchase_order=purchase_order, expenses=expenses, recurring_expenses=recurring_expenses,
                payment_made=payment_made,
                projects=projects,
                chart_of_accounts=chart_of_accounts, manual_journal=manual_journal,
                eway_bill=eway_bill,
                employees=employees, employees_loan=employees_loan, holiday=holiday,
                attendance=attendance, salary_details=salary_details,
                reports=reports,update_action=update_action,status=status    
            )
            data.save()
            messages.success(request,"Request sent successfully. Please wait for approval.")
            return redirect('company_profile')
        else:
            return redirect('company_module_editpage')  
    else:
        return redirect('/')


def company_renew_terms(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        # Check for any previous  extension request
        if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists():
            messages.warning(request,'You have a pending request, wait for approval or contact our support team for any help..?')
            return redirect('company_profile')
        if request.method == 'POST':
            select=request.POST['select']
            terms=PaymentTerms.objects.get(id=select)
            update_action=1
            status='Pending'
            newterms=PaymentTermsUpdates(
               company=dash_details,
               payment_term=terms,
               update_action=update_action,
               status=status 
            )
            newterms.save()
            messages.success(request,'Request sent successfully, Please wait for approval...')
            return redirect('company_profile')
        else:
            return redirect('company_profile')
    else:
        return redirect('/')

# company notifications and messages
def company_notifications(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        notifications = dash_details.notifications.filter(is_read=0).order_by('-date_created','-time')
        end_date = dash_details.End_date
        company_days_remaining = (end_date - date.today()).days
        payment_request = True if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists() else False
        
        print(company_days_remaining)
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'notifications':notifications,
            'days_remaining':company_days_remaining,
            'payment_request':payment_request,
        }

        return render(request,'company/company_notifications.html',context)
        
    else:
        return redirect('/')
        
        
def company_message_read(request,pk):
    '''
    message read functions set the is_read to 1, 
    by default it is 0 means not seen by user.

    '''
    notification=Notifications.objects.get(id=pk)
    notification.is_read=1
    notification.save()
    return redirect('company_notifications')
    
    
def company_payment_history(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/') 
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        payment_history=dash_details.previous_plans.all()

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'payment_history':payment_history,
            
        }
        return render(request,'company/company_payment_history.html', context)
    else:
        return redirect('/')
        
def company_trial_feedback(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/') 
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        trial_instance = TrialPeriod.objects.get(company=dash_details)
        if request.method == 'POST':
            interested = request.POST.get('interested')
            feedback=request.POST.get('feedback') 
            
            trial_instance.interested_in_buying=1 if interested =='yes' else 2
            trial_instance.feedback=feedback
            trial_instance.save()

            if interested =='yes':
                return redirect('company_profile')
            else:
                return redirect('company_dashboard')
        else:
            return redirect('company_dashboard')
    else:
        return redirect('/')
# -------------------------------Staff section--------------------------------

# staff dashboard
def staff_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_dash.html',context)
    else:
        return redirect('/')


# staff profile
def staff_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_profile.html',context)
    else:
        return redirect('/')


def staff_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'staff/staff_profile_editpage.html', context)
    else:
        return redirect('/')

def staff_profile_details_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            dash_details.contact = request.POST.get('phone')
            old=dash_details.image
            new=request.FILES.get('profile_pic')
            print(new,old)
            if old!=None and new==None:
                dash_details.image=old
            else:
                print(new)
                dash_details.image=new
            dash_details.save()
            messages.success(request,'Updated')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')

def staff_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                if LoginDetails.objects.filter(password=password).exists():
                    messages.error(request,'Use another password')
                    return redirect('staff_profile_editpage')
                else:
                    log_details.password=password
                    log_details.save()

            messages.success(request,'Password Changed')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')


    
def company_gsttype_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            
            gstno = request.POST.get('gstno')
            gsttype = request.POST.get('gsttype')

            # Check if gsttype is one of the specified values
            if gsttype in ['unregistered Business', 'Overseas', 'Consumer']:
                dash_details.gst_no = None
            else:
                if gstno:
                    dash_details.gst_no = gstno
                else:
                    messages.error(request,'GST Number is not entered*')
                    return redirect('company_profile_editpage')


            dash_details.gst_type = gsttype

            dash_details.save()
            messages.success(request,'GST Type changed')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/') 
    

# -------------------------------Zoho Modules section--------------------------------

#--------------Customer-----------------#
#-------------------Arya E.R----------------------#

def customer(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
        
        comp_payment_terms=Company_Payment_Term.objects.filter(company=comp_details)
        price_lists=PriceList.objects.filter(company=comp_details,type='Sales',status='Active')

       
        return render(request,'zohomodules/customer/create_customer.html',{'details':dash_details,'allmodules': allmodules,'comp_payment_terms':comp_payment_terms,'log_details':log_details,'price_lists':price_lists}) 
    else:
        return redirect('/')  

def view_customer_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')

        data=Customer.objects.filter(company=comp_details)

        

        return render(request,'zohomodules/customer/customer_list.html',{'details':dash_details,'allmodules': allmodules,'data':data,'log_details':log_details}) 


    else:
        return redirect('/')
    

def add_customer(request):
   
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')

        

       
        if request.method=="POST":
            customer_data=Customer()
            customer_data.login_details=log_details
            customer_data.company=comp_details
            customer_data.customer_type = request.POST.get('type')

            customer_data.title = request.POST.get('salutation')
            customer_data.first_name=request.POST['first_name']
            customer_data.last_name=request.POST['last_name']
            customer_data.company_name=request.POST['company_name']
            customer_data.customer_display_name=request.POST['v_display_name']
            customer_data.customer_email=request.POST['vendor_email']
            customer_data.customer_phone=request.POST['w_phone']
            customer_data.customer_mobile=request.POST['m_phone']
            customer_data.skype=request.POST['skype_number']
            customer_data.designation=request.POST['designation']
            customer_data.department=request.POST['department']
            customer_data.website=request.POST['website']
            customer_data.GST_treatement=request.POST['gst']
            customer_data.customer_status="Active"
            customer_data.remarks=request.POST['remark']
            customer_data.current_balance=request.POST['opening_bal']

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                customer_data.PAN_number=request.POST['pan_number']
                customer_data.GST_number="null"
            else:
                customer_data.GST_number=request.POST['gst_number']
                customer_data.PAN_number=request.POST['pan_number']

            customer_data.place_of_supply=request.POST['source_supply']
            customer_data.currency=request.POST['currency']
            op_type=request.POST.get('op_type')
            if op_type is not None:
                customer_data.opening_balance_type=op_type
            else:
                customer_data.opening_balance_type='Opening Balance not selected'
    
            customer_data.opening_balance=request.POST['opening_bal']
            customer_data.company_payment_terms=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])
            # customer_data.price_list=request.POST['plst']
            plst=request.POST.get('plst')
            if plst!=0:
                 customer_data.price_list=plst
            else:
                customer_data.price_list='Price list not selected'




            # customer_data.portal_language=request.POST['plang']
            plang=request.POST.get('plang')
            if plang!=0:
                 customer_data.portal_language=plang
            else:
                customer_data.portal_language='Portal language not selected'

            customer_data.facebook=request.POST['fbk']
            customer_data.twitter=request.POST['twtr']
            customer_data.tax_preference=request.POST['tax1']

            type=request.POST.get('type')
            if type is not None:
                customer_data.customer_type=type
            else:
                customer_data.customer_type='Customer type not selected'
    



           
            customer_data.billing_attention=request.POST['battention']
            customer_data.billing_country=request.POST['bcountry']
            customer_data.billing_address=request.POST['baddress']
            customer_data.billing_city=request.POST['bcity']
            customer_data.billing_state=request.POST['bstate']
            customer_data.billing_pincode=request.POST['bzip']
            customer_data.billing_mobile=request.POST['bphone']
            customer_data.billing_fax=request.POST['bfax']
            customer_data.shipping_attention=request.POST['sattention']
            customer_data.shipping_country=request.POST['s_country']
            customer_data.shipping_address=request.POST['saddress']
            customer_data.shipping_city=request.POST['scity']
            customer_data.shipping_state=request.POST['sstate']
            customer_data.shipping_pincode=request.POST['szip']
            customer_data.shipping_mobile=request.POST['sphone']
            customer_data.shipping_fax=request.POST['sfax']
            customer_data.save()
           # ................ Adding to History table...........................
            
            vendor_history_obj=CustomerHistory()
            vendor_history_obj.company=comp_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.customer=customer_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Completed'
            vendor_history_obj.save()

    # .......................................................adding to remaks table.....................
            vdata=Customer.objects.get(id=customer_data.id)
            vendor=vdata
            rdata=Customer_remarks_table()
            rdata.remarks=request.POST['remark']
            rdata.company=comp_details
            rdata.customer=vdata
            rdata.save()


     #...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            vdata=Customer.objects.get(id=customer_data.id)
            vendor=vdata
           
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department)
                    mapped2=list(mapped2)
                    print(mapped2)
                    for ele in mapped2:
                        created = CustomerContactPersons.objects.get_or_create(title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype=ele[6],designation=ele[7],department=ele[8],company=comp_details,customer=vendor)
                
        
            messages.success(request, 'Customer created successfully!')   

            return redirect('view_customer_list')
        
        else:
            messages.error(request, 'Some error occurred !')   

            return redirect('view_customer_list')





def check_customer_phonenumber_exist(request):
    if request.method == 'GET':
       mPhone = request.GET.get('m_Phone', None)

       if mPhone:
          
            exists = Customer.objects.filter(
                    customer_mobile=mPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False}) 

def check_customer_work_phone_exist(request):
    if request.method == 'GET':
       wPhone = request.GET.get('w_Phone', None)

       if wPhone:
          
            exists = Customer.objects.filter(
                    customer_phone=wPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})   

def check_customer_email_exist(request):
    if request.method == 'GET':
       vendoremail = request.GET.get('vendor_email', None)

       if vendoremail:
          
            exists = Customer.objects.filter(
                    customer_email=vendoremail
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False}) 

def customer_payment_terms_add(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
    
        if request.method == 'POST':
            terms = request.POST.get('name')
            day = request.POST.get('days')
            normalized_data = terms.replace(" ", "")
            pay_tm = add_space_before_first_digit(normalized_data)
            ptr = Company_Payment_Term(term_name=pay_tm, days=day, company=dash_details)
            ptr.save()
            payterms_obj = Company_Payment_Term.objects.filter(company=comp_details).values('id', 'term_name')


            payment_list = [{'id': pay_terms['id'], 'name': pay_terms['term_name']} for pay_terms in payterms_obj]
            response_data = {
            "message": "success",
            'payment_list':payment_list,
            }
            return JsonResponse(response_data)

        else:
            return JsonResponse({'error': 'Invalid request'}, status=400)   
            
def add_space_before_first_digit(data):
    for index, char in enumerate(data):
        if char.isdigit():
            return data[:index] + ' ' + data[index:]
    return data





def check_customer_term_exist(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

    if request.method == 'GET':
       term_name = request.GET.get('term_name', None)
       if term_name:
            normalized_data = term_name.replace(" ", "")
            term_name_processed = add_space_before_first_digit(normalized_data)
            exists = Company_Payment_Term.objects.filter(
                    term_name=term_name_processed,
                    company=comp_details
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})    

def customer_check_pan(request):
    if request.method == 'POST':
        panNumber = request.POST.get('panNumber')
        pan_exists = Customer.objects.filter(PAN_number=panNumber).exists()

        if pan_exists:
            return JsonResponse({'status': 'exists'})
        else:
            return JsonResponse({'status': 'not_exists'})
    else:
        return JsonResponse({'error': 'Invalid request'})  

def customer_check_gst(request):
    if request.method == 'POST':
        gstNumber = request.POST.get('gstNumber')
        gst_exists = Customer.objects.filter(GST_number=gstNumber).exists()
       
        if gst_exists:
            return JsonResponse({'status': 'exists'})
        else:
            return JsonResponse({'status': 'not_exists'})
    else:
        return JsonResponse({'error': 'Invalid request'}) 

def sort_customer_by_name(request):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
  
        data=Customer.objects.filter(login_details=log_details).order_by('first_name')
        return render(request,'zohomodules/customer/customer_list.html',{'data':data,'dash_details':dash_details})
     else:
            return redirect('/')    

def sort_customer_by_amount(request):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
        data=Customer.objects.filter(login_details=log_details).order_by('opening_balance')
        return render(request,'zohomodules/customer/customer_list.html',{'data':data,'dash_details':dash_details})
     else:
         return redirect('/')   


def view_customer_active(request):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
        data=Customer.objects.filter(login_details=log_details,customer_status='Active').order_by('-id')
        return render(request,'zohomodules/customer/customer_list.html',{'data':data,'dash_details':dash_details})
     else:
         return redirect('/') 

    
    
def view_customer_inactive(request):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   
        data=Customer.objects.filter(login_details=log_details,customer_status='Inactive').order_by('-id')
        return render(request,'zohomodules/customer/customer_list.html',{'data':data,'dash_details':dash_details})
     else:
         return redirect('/') 


def import_customer_excel(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   
        if request.method == 'POST' :
       
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(2, eb.max_row + 1):
                            
                    vendorsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    comp_term=vendorsheet[18]
                    if comp_term:
                        normalized_data = comp_term.replace(" ", "")

                        pay_tm = add_space_before_first_digit(normalized_data)
                    else:
                        cpt =Company_Payment_Term.objects.filter(company=comp_details).first()
                        pay_tm = cpt.term_name
   
                    try:
                        com_term_obj=Company_Payment_Term.objects.get(company=comp_details,term_name=pay_tm)
                    except Company_Payment_Term.DoesNotExist:
                        com_term_obj= None
                    
                    opn_blc_str = vendorsheet[17]  # Assuming vendorsheet[15] is a string representing a decimal
                    if opn_blc_str:

                        opn_blc = float(opn_blc_str)
                    else:
                        opn_blc = 0.00    
                    

                    Vendor_object=Customer(customer_type=vendorsheet[0],title=vendorsheet[1],first_name=vendorsheet[2],last_name=vendorsheet[3],company_name=vendorsheet[4],customer_email=vendorsheet[5],customer_phone=vendorsheet[6],customer_mobile=vendorsheet[7],skype=vendorsheet[8],designation=vendorsheet[9],department=vendorsheet[10],website=vendorsheet[11],
                                         GST_treatement=vendorsheet[12],place_of_supply=vendorsheet[13],tax_preference=vendorsheet[14],currency=vendorsheet[15],opening_balance_type=vendorsheet[16],
                                         opening_balance=opn_blc,company_payment_terms=com_term_obj,billing_attention=vendorsheet[19],billing_country=vendorsheet[20],billing_address=vendorsheet[21],
                                         billing_city=vendorsheet[22],billing_state=vendorsheet[23],billing_pincode=vendorsheet[24],
                                         billing_mobile=vendorsheet[25],billing_fax=vendorsheet[26],shipping_attention=vendorsheet[27],shipping_country=vendorsheet[28],shipping_address=vendorsheet[29],shipping_city=vendorsheet[30],
                                         shipping_state=vendorsheet[31],shipping_pincode=vendorsheet[32],
                                         shipping_mobile=vendorsheet[33], shipping_fax=vendorsheet[34], remarks=vendorsheet[35],current_balance=opn_blc,customer_status="Active",company=comp_details,login_details=log_details)
                    Vendor_object.save()

    
                   
                messages.warning(request,'file imported')
                return redirect('view_customer_list')    

    
            messages.error(request,'File upload Failed!11')
            return redirect('view_customer_list')
        else:
            messages.error(request,'File upload Failed!11')
            return redirect('view_customer_list') 

def view_customer_details(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

        vendor_obj=Customer.objects.get(id=pk)

        # Getting all vendor to disply on the left side of vendor_detailsnew page
        vendor_objs=Customer.objects.filter(company=comp_details)

        vendor_comments=Customer_comments_table.objects.filter(customer=vendor_obj)
        vendor_history=CustomerHistory.objects.filter(customer=vendor_obj)
    
    content = {
                'details': dash_details,
               
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
                'log_details':log_details,
                'vendor_objs':vendor_objs,
                'vendor_comments':vendor_comments,
                'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/customer/customer_detailsnew.html',content)    

def sort_customer(request,selectId,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

        vendor_obj = Customer.objects.get(id=pk)
        vendor_objs = Customer.objects.filter(company=comp_details)

        if selectId == 0:
            vendor_objs=Customer.objects.filter(company=comp_details)
        if selectId == 1:
            vendor_objs=Customer.objects.filter(company=comp_details).order_by('first_name')
        if selectId == 2:
            vendor_objs=Customer.objects.filter(company=comp_details).order_by('opening_balance')
           
        
        vendor_comments=Customer_comments_table.objects.filter(customer=vendor_obj)
        vendor_history=CustomerHistory.objects.filter(customer=vendor_obj)
    
        content = {
                'details': dash_details,
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
                'log_details':log_details,
                'vendor_objs':vendor_objs,
                'vendor_comments':vendor_comments,
                'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/customer/customer_detailsnew.html',content) 

def customer_status_change(request,statusId,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   
    

        vendor_obj = Customer.objects.get(id=pk)
        vendor_objs = Customer.objects.filter(company=comp_details)

        if statusId == 0:
            vendor_objs=Customer.objects.filter(company=comp_details)
        if statusId == 1:
            vendor_objs=Customer.objects.filter(company=comp_details,customer_status='Active').order_by('-id')
        if statusId == 2:
            vendor_objs=Customer.objects.filter(company=comp_details,customer_status='Inactive').order_by('-id')
           
        
        vendor_comments=Customer_comments_table.objects.filter(customer=vendor_obj)
        vendor_history=CustomerHistory.objects.filter(customer=vendor_obj)
    
        content = {
                'details': dash_details,
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
                'log_details':log_details,
                'vendor_objs':vendor_objs,
                'vendor_comments':vendor_comments,
                'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/customer/customer_detailsnew.html',content)       

def delete_customers(request, pk):
    try:
        vendor_obj = Customer.objects.get(id=pk)

        vendor_obj.delete()
        return redirect('view_customer_list')  
    except Customer.DoesNotExist:
        return HttpResponseNotFound("Customer not found.")  

def customer_status(request,pk):
    vendor_obj = Customer.objects.get(id=pk)
    if vendor_obj.customer_status == 'Active':
        vendor_obj.customer_status ='Inactive'
    elif vendor_obj.customer_status == 'Inactive':
        vendor_obj.customer_status ='Active'
    vendor_obj.save()
    return redirect('view_customer_details',pk)         

def customer_add_comment(request,pk):
   if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   
        if request.method =='POST':
            comment_data=request.POST['comments']
       
            vendor_id= Customer.objects.get(id=pk) 
            vendor_obj=Customer_comments_table()
            vendor_obj.comment=comment_data
            vendor_obj.customer=vendor_id
            vendor_obj.company=comp_details
            vendor_obj.login_details= LoginDetails.objects.get(id=log_id)

            vendor_obj.save()
            return redirect('view_customer_details',pk)
   return redirect('view_customer_details',pk) 


def customer_delete_comment(request, pk):
    try:
        vendor_comment =Customer_comments_table.objects.get(id=pk)
        vendor_id=vendor_comment.customer.id
        vendor_comment.delete()
        return redirect('view_customer_details',vendor_id)  
    except Customer_comments_table.DoesNotExist:
        return HttpResponseNotFound("comments not found.") 

def add_customer_file(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

        if request.method == 'POST':
            data=request.FILES.getlist('file')
            try:
                for doc in data:

                    vendor_obj=Customer_doc_upload_table()
                    
                    vendor_obj.document = doc
                    vendor_obj.login_details = log_details
                    vendor_obj.company = comp_details
                    vendor_obj.customer = Customer.objects.get(id=pk)
                    vendor_obj.save()
                
                messages.success(request,'File uploaded')
                return redirect('view_customer_details',pk) 
            except Customer_doc_upload_table.DoesNotExist:
                return redirect('view_customer_details',pk) 

 
            

def customer_shareemail(request,pk):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
    
    
        vendor_obj=Customer.objects.get(id=pk)

        context = {'vendor_obj':vendor_obj,'details':dash_details}
        if request.method == 'POST':
            try:
                emails_string = request.POST['email_ids']

                        # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                                                                                          
                template_path = 'zohomodules/customer/customermailoverview.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                subject = f"Transaction Details"
                email = f"Hi,\nPlease find the attached transaction details {vendor_obj.first_name} {vendor_obj.last_name}.\n"
                email_from = settings.EMAIL_HOST_USER

        
                msg = EmailMultiAlternatives(subject, email, email_from, emails_list)
                msg.attach(f'{vendor_obj.first_name}_{vendor_obj.last_name}_Transactions.pdf', pdf, "application/pdf")
                
                # Send the email
                msg.send()

                messages.success(request, 'Transaction has been shared via email successfully..!')
                return redirect('view_customer_details',pk)

            except Exception as e:
                print(f"Error sending email: {e}")
                messages.error(request, 'An error occurred while sending the email. Please try again later.')
                return redirect('view_customer_details',pk)





def Customer_edit(request,pk):
   
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')

        customer_obj=Customer.objects.get(id=pk)

        customer_contact_obj=CustomerContactPersons.objects.filter(customer=customer_obj)  
        comp_payment_terms=Company_Payment_Term.objects.filter(company=comp_details)
        price_lists=PriceList.objects.filter(company=comp_details,type='Sales',status='Active')
    
        content = {
                'details': dash_details,
                'allmodules': allmodules,
                'customer_obj':customer_obj,
                'log_details':log_details,
                'customer_contact_obj':customer_contact_obj,
                'comp_payment_terms':comp_payment_terms,
                'price_lists': price_lists,
        }
    

        return render(request,'zohomodules/customer/edit_customer.html',content)
     else:
            return redirect('/')

def do_customer_edit(request,pk):
         
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
    
        if request.method=="POST":
            customer_data=Customer.objects.get(id=pk)
            customer_data.login_details=log_details
            customer_data.company=comp_details
            customer_data.title = request.POST.get('salutation')
            customer_data.first_name=request.POST['first_name']
            customer_data.last_name=request.POST['last_name']
            customer_data.company_name=request.POST['company_name']
            customer_data.customer_display_name=request.POST['v_display_name']
            customer_data.customer_email=request.POST['vendor_email']
            customer_data.customer_phone=request.POST['w_phone']
            customer_data.customer_mobile=request.POST['m_phone']
            customer_data.skype=request.POST['skype_number']
            customer_data.designation=request.POST['designation']
            customer_data.department=request.POST['department']
            customer_data.website=request.POST['website']
            customer_data.GST_treatement=request.POST['gst']
            customer_data.customer_status="Active"
            customer_data.remarks=request.POST['remark']
            
            cob=float(request.POST['opening_bal'])
            oc=float(customer_data.current_balance) 
            ob=float(customer_data.opening_balance) 


            if cob > ob:
                diffadd=cob-ob
                oc=oc + diffadd
                customer_data.current_balance=oc
                customer_data.opening_balance=cob
            elif cob < ob:
                diffadd=ob-cob
                oc=oc-diffadd
                customer_data.current_balance=oc
                customer_data.opening_balance=cob

            else:
                customer_data.current_balance=request.POST['opening_bal']   
       
            

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                customer_data.PAN_number=request.POST['pan_number']
                customer_data.GST_number="null"
            else:
                customer_data.GST_number=request.POST['gst_number']
                customer_data.PAN_number=request.POST['pan_number']

            customer_data.place_of_supply=request.POST['source_supply']
            customer_data.currency=request.POST['currency']
            op_type=request.POST.get('op_type')
            if op_type is not None:
                customer_data.opening_balance_type=op_type
            else:
                customer_data.opening_balance_type='Opening Balance not selected'
            customer_data.opening_balance=request.POST['opening_bal']
            customer_data.company_payment_terms=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])
            plst=request.POST.get('plst')
            if plst!=0:

                 customer_data.price_list=plst
            else:
                customer_data.price_list='Price list not selected'


            # customer_data.portal_language=request.POST['plang']
            plang=request.POST.get('plang')
            if plang!=0:
                 customer_data.portal_language=plang
            else:
                customer_data.portal_language='Portal language not selected'

            customer_data.facebook=request.POST['fbk']
            customer_data.twitter=request.POST['twtr']
            customer_data.tax_preference=request.POST['tax1']

            type=request.POST.get('type')
            if type is not None:
                customer_data.customer_type=type
            else:
                customer_data.customer_type='Customer type not selected'
            
           
            customer_data.billing_attention=request.POST['battention']
            customer_data.billing_country=request.POST['bcountry']
            customer_data.billing_address=request.POST['baddress']
            customer_data.billing_city=request.POST['bcity']
            customer_data.billing_state=request.POST['bstate']
            customer_data.billing_pincode=request.POST['bzip']
            customer_data.billing_mobile=request.POST['bphone']
            customer_data.billing_fax=request.POST['bfax']
            customer_data.shipping_attention=request.POST['sattention']
            customer_data.shipping_country=request.POST['s_country']
            customer_data.shipping_address=request.POST['saddress']
            customer_data.shipping_city=request.POST['scity']
            customer_data.shipping_state=request.POST['sstate']
            customer_data.shipping_pincode=request.POST['szip']
            customer_data.shipping_mobile=request.POST['sphone']
            customer_data.shipping_fax=request.POST['sfax']
            customer_data.save()


              # ................ Adding to History table...........................
            
            vendor_history_obj=CustomerHistory()
            vendor_history_obj.company=comp_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.customer=customer_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Edited'
            vendor_history_obj.save()
    # .......................................................adding to remaks table.....................
            vdata=Customer.objects.get(id=customer_data.id)
            try:

                rdata=Customer_remarks_table.objects.get(customer=vdata)
                rdata.remarks=request.POST['remark']
                rdata.company=comp_details
                rdata.customer=vdata
                rdata.save()
            except Customer_remarks_table.DoesNotExist:
                remarks_obj= Customer_remarks_table()   
                remarks_obj.remarks=request.POST['remark']
                remarks_obj.company=comp_details
                remarks_obj.customer=vdata
                remarks_obj.save()


    #  ...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            person = request.POST.getlist('contact_person_id[]')
            vdata=Customer.objects.get(id=customer_data.id)
            vendor=vdata
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department)==len(person):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department,person)
                    mapped2=list(mapped2)
                    for ele in mapped2:
                       
                        existing_instance = CustomerContactPersons.objects.filter(id=ele[9], company=comp_details, customer=vendor).first()
                        if existing_instance:
                            # Update the existing instance
                            existing_instance.title = ele[0]
                            existing_instance.first_name = ele[1]
                            existing_instance.last_name = ele[2]
                            existing_instance.email = ele[3]
                            existing_instance.work_phone  = ele[4]
                            existing_instance.mobile = ele[5]
                            existing_instance.skype = ele[6]
                            existing_instance.designation = ele[7]
                            existing_instance.department = ele[8]

                            # Update other fields

                            existing_instance.save()
                        else:
                            # Create a new instance
                            new_instance = CustomerContactPersons.objects.create(
                                title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype_name_number=ele[6],designation=ele[7],department=ele[8],company=comp_details,customer=vendor
                            )
            return redirect('view_customer_details',pk)  

#------------------------------------End----------------------------------------------#                                             